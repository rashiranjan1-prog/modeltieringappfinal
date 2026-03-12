

# VERSION: 2026-03-12-v6 (METADATA_COLS=7 fix, robust weight matching, hidden cols included)
"""
Excel loader — supports two formats:

Format A (standard):  Sheets named Models, Parameters, Tiers, Settings, Scores
Format B (matrix):    Sheets named Tiering_Method, Parameters, Model_tier, Models_List
                      (the wide pivot format where models are columns and scores are cells)

Improvements:
- Fuzzy model name matching (strips whitespace, case-insensitive)
- Param name matching is also normalised
- Param weights from Tiering_Method ALWAYS override the default
- Robust Internal / Tier row detection
"""
from openpyxl import load_workbook
from ..db import get_db
from .tieringlogic import compute_tiering_for_all


def _sheet(wb, *names):
    lookup = {s.lower(): s for s in wb.sheetnames}
    for name in names:
        if name.lower() in lookup:
            return wb[lookup[name.lower()]]
    return None


def _headers(row):
    return [str(c).strip().lower().replace(' ', '_').replace('-', '_') if c else '' for c in row]


def _parse_weight(val):
    """Parse weight: handles 0.35, '35%', 35 → float 0.0–1.0"""
    if val is None:
        return None
    try:
        f = float(val)
        return f / 100.0 if f > 1.0 else f
    except (ValueError, TypeError):
        s = str(val).strip().replace('%', '')
        try:
            return float(s) / 100.0
        except ValueError:
            return None


def _norm(s):
    """Normalise a string for fuzzy matching."""
    return str(s or '').strip().lower()


def load_excel(filepath):
    wb = load_workbook(filepath, data_only=True)
    results = {'models': 0, 'parameters': 0, 'tiers': 0, 'settings': 0,
               'scores': 0, 'computed': 0}

    sheet_names_lower = [s.lower() for s in wb.sheetnames]
    is_matrix = 'tiering_method' in sheet_names_lower or 'model_tier' in sheet_names_lower

    if is_matrix:
        _load_matrix_format(wb, filepath, results)
    else:
        _load_standard_format(wb, results)

    results['computed'] = _smart_compute(results.get('score_row_found', False))
    return results


def _smart_compute(score_row_found):
    """
    Compute tiering for all models.
    If the Excel Internal row was loaded, trust those scores for models that have them
    and only recompute models that have no computed_score yet (the new ones).
    """
    from .tieringlogic import compute_tiering_for_model
    db = get_db()

    if not score_row_found:
        # No Internal row scores saved — recompute everything from param scores
        from .tieringlogic import compute_tiering_for_all
        return compute_tiering_for_all()

    models = db.execute('SELECT * FROM models').fetchall()
    tiers  = db.execute('SELECT * FROM tiers ORDER BY sort_order').fetchall()
    count  = 0

    for m in models:
        mid   = m['id']
        score = m['computed_score']

        if score and float(score) > 0:
            # Trust the Excel Internal row score — just assign the tier
            from .tieringlogic import _match_tier
            matched = _match_tier(float(score), tiers)
            # Respect any existing manual override
            was_overridden = (m['current_tier'] and m['computed_tier'] and
                              m['current_tier'] != m['computed_tier'])
            current = m['current_tier'] if was_overridden else matched
            db.execute(
                'UPDATE models SET computed_tier=?, current_tier=?,'
                ' last_computed_at=datetime("now") WHERE id=?',
                (matched, current, mid)
            )
        else:
            # No Internal row score for this model — recompute from param scores
            compute_tiering_for_model(mid)

        count += 1

    db.commit()
    return count


def _get_hidden_col_indices(filepath, sheet_name):
    hidden = set()
    try:
        wb2 = load_workbook(filepath, data_only=False)
        for sname in wb2.sheetnames:
            if sname.lower() == sheet_name.lower():
                ws = wb2[sname]
                from openpyxl.utils import column_index_from_string
                for col_letter, dim in ws.column_dimensions.items():
                    if dim.hidden:
                        hidden.add(column_index_from_string(col_letter))
                break
    except Exception:
        pass
    return hidden


def _load_matrix_format(wb, filepath, results):
    db = get_db()

    # Full wipe for clean reimport
    import sys
    print("\n\n====== EXCELLOADER v5 RUNNING ======", file=sys.stderr, flush=True)
    db.execute('DELETE FROM model_scores')
    db.execute('DELETE FROM overrides')
    db.execute('DELETE FROM models')
    db.execute('DELETE FROM parameters')
    db.execute("DELETE FROM sqlite_sequence WHERE name IN "
               "('models','parameters','model_scores','overrides')")
    db.commit()

    hidden_cols = _get_hidden_col_indices(filepath, 'Tiering_Method')

    # ── 1. Parameters sheet ──────────────────────────────────────────────────
    param_map      = {}   # exact sub_parameter name → param_id
    param_map_norm = {}   # normalised name → param_id  (for fuzzy matching)

    ws_params = _sheet(wb, 'Parameters')
    if ws_params:
        rows = list(ws_params.iter_rows(values_only=True))
        current_group = None
        for row in rows[1:]:
            if not any(row):
                continue
            grp_val = str(row[0]).strip() if row[0] else None
            sub     = str(row[1]).strip() if len(row) > 1 and row[1] else None
            desc    = str(row[2]).strip() if len(row) > 2 and row[2] else ''
            l1      = str(row[3]).strip() if len(row) > 3 and row[3] else ''
            l2      = str(row[4]).strip() if len(row) > 4 and row[4] else ''
            l3      = str(row[5]).strip() if len(row) > 5 and row[5] else ''

            if grp_val and grp_val.lower() not in ('none', 'parameter'):
                current_group = grp_val
            if not sub or not current_group:
                continue
            if sub.lower() in ('none', 'sub-parameter', 'sub_parameter'):
                continue

            c = db.execute(
                'INSERT INTO parameters '
                '(grp, sub_parameter, criteria, description, weight,'
                ' level1_label, level2_label, level3_label) VALUES (?,?,?,?,?,?,?,?)',
                (current_group, sub, '', desc, 1.0,
                 l1 or 'Low', l2 or 'Medium', l3 or 'High')
            )
            pid = c.lastrowid
            param_map[sub]         = pid
            param_map_norm[_norm(sub)] = pid
            results['parameters'] += 1
        db.commit()

    # ── 2. Models (Model_tier sheet) ─────────────────────────────────────────
    model_map      = {}   # exact name → model_id
    model_map_norm = {}   # normalised name → model_id

    ws_models = _sheet(wb, 'Model_tier', 'Model_List', 'Models_List', 'Models')
    if ws_models:
        rows = list(ws_models.iter_rows(values_only=True))
        hdrs = _headers(rows[0])
        name_col = next((i for i, h in enumerate(hdrs) if 'name' in h or 'model' in h), 1)
        risk_col = next((i for i, h in enumerate(hdrs) if 'risk' in h), 2)
        tier_col = next((i for i, h in enumerate(hdrs) if 'tier' in h), None)
        for row in rows[1:]:
            if not any(row):
                continue
            name = str(row[name_col]).strip() if len(row) > name_col and row[name_col] else None
            risk = str(row[risk_col]).strip() if len(row) > risk_col and row[risk_col] else ''
            tier = (str(row[tier_col]).strip()
                    if tier_col is not None and len(row) > tier_col and row[tier_col] else None)
            if not name or name.lower() in ('none', '', 'model name', 'model_name'):
                continue
            c = db.execute('INSERT INTO models (name, risk_type, current_tier) VALUES (?,?,?)',
                           (name, risk, tier))
            mid = c.lastrowid
            model_map[name]         = mid
            model_map_norm[_norm(name)] = mid
            results['models'] += 1
        db.commit()

    # ── 3. Tiering_Method matrix ─────────────────────────────────────────────
    ws_matrix = _sheet(wb, 'Tiering_Method', 'Tiering Method')
    if not ws_matrix or not model_map or not param_map:
        return

    rows = list(ws_matrix.iter_rows(values_only=True))
    if not rows:
        return

    header = rows[0]

    # Find MODEL_START_COL: first visible col whose header is a known model name
    def _find_model_id(cell_val):
        """Return model_id for a header cell, using exact then fuzzy match."""
        if not cell_val:
            return None
        s = str(cell_val).strip()
        if s in model_map:
            return model_map[s]
        ns = _norm(s)
        return model_map_norm.get(ns)

    # Cols 1-7 (A-G) are ALWAYS metadata: Parameter, Sub-Param, Description,
    # Value Range, Weight, duplicate Weight, spacer. Never model columns.
    METADATA_COLS = 7

    MODEL_START_COL = None
    for col_idx, cell_val in enumerate(header, 1):
        if col_idx <= METADATA_COLS:   # always skip metadata cols
            continue
        if col_idx in hidden_cols:
            continue
        if _find_model_id(cell_val) is not None:
            MODEL_START_COL = col_idx
            break
    if MODEL_START_COL is None:
        # fallback ignoring hidden flag but still respect metadata boundary
        for col_idx, cell_val in enumerate(header, 1):
            if col_idx <= METADATA_COLS:
                continue
            if _find_model_id(cell_val) is not None:
                MODEL_START_COL = col_idx
                break
    if MODEL_START_COL is None:
        MODEL_START_COL = 8

    # Build list of (col_index, model_id) for all model columns
    # Skip cols 1-7 (metadata) but INCLUDE hidden cols — hidden model columns still have scores
    model_cols = []
    for col_idx in range(MODEL_START_COL, len(header) + 1):
        if col_idx <= METADATA_COLS:
            continue
        # Do NOT skip hidden cols — users hide model columns in Excel but scores still exist
        cell_val = header[col_idx - 1] if col_idx - 1 < len(header) else None
        mid = _find_model_id(cell_val)
        if mid is not None:
            model_cols.append((col_idx, mid))

    WEIGHT_COL = 5   # col E
    SUB_COL    = 2   # col B
    GRP_COL    = 1   # col A

    group_weight_rows = {}
    param_weight_rows = {}   # normalised sub_param → weight
    param_score_rows  = {}   # normalised sub_param → row idx
    internal_score_row = None
    tier_table_start   = None

    SKIP_KEYWORDS = {'override', 'mrm', 'tier assignment'}

    for idx, row in enumerate(rows):
        if not any(row):
            continue
        grp_val = str(row[GRP_COL - 1] or '').strip() if len(row) >= GRP_COL else ''
        sub_val = str(row[SUB_COL - 1] or '').strip() if len(row) >= SUB_COL else ''

        # Detect Internal score row:
        # Try text label first (col A-D), then fall back to float-value detection below
        any_col_vals = [str(row[c] or '').strip().lower() for c in range(min(4, len(row)))]
        if any(v in ('internal', 'internal score') or v.startswith('internal') for v in any_col_vals if v):
            internal_score_row = idx
            continue
        if grp_val.lower() == 'tier' and sub_val.lower() in ('', 'lower range', 'lower_range'):
            tier_table_start = idx
            continue
        if grp_val.lower() in SKIP_KEYWORDS:
            continue

        if grp_val in ('Materiality', 'Criticality', 'Complexity'):
            w = _parse_weight(row[WEIGHT_COL - 1] if len(row) >= WEIGHT_COL else None)
            if w is not None:
                group_weight_rows[grp_val] = w
            continue

        if sub_val and sub_val.lower() not in ('none', '', 'sub-parameter'):
            w   = _parse_weight(row[WEIGHT_COL - 1] if len(row) >= WEIGHT_COL else None)
            nk  = _norm(sub_val)
            if w is not None:
                param_weight_rows[nk] = w
            # Record score row if this param exists in param_map (exact or fuzzy)
            if sub_val in param_map or nk in param_map_norm:
                param_score_rows[nk] = idx

    # --- Fallback: find Internal score row by float values in model columns ---
    import sys
    print(f"====== internal_score_row after text scan: {internal_score_row} ======", file=sys.stderr, flush=True)
    print(f"====== model_cols count: {len(model_cols)} ======", file=sys.stderr, flush=True)
    # Handles case where 'Internal' is a floating textbox (not a real cell value)
    if internal_score_row is None and model_cols:
        # Look for a row after row 18 where model columns contain float scores (not 1/2/3)
        # Param rows only have integers 1, 2, 3. Score row has floats like 1.86.
        last_known_param_idx = max(
            (idx for idx, row in enumerate(rows)
             if not any(row) is False
             and any(row[c-1] if c-1 < len(row) else None
                     for c in [r for _, r in [(0, param_score_rows[k]) for k in param_score_rows]])
             ), default=18
        ) if param_score_rows else 18
        # Simpler: scan rows 19 onwards for float values
        for idx in range(18, min(30, len(rows))):
            row = rows[idx]
            model_vals = [row[col_idx-1] for col_idx, _ in model_cols[:5]
                          if col_idx-1 < len(row) and row[col_idx-1] is not None]
            float_vals = [v for v in model_vals
                          if isinstance(v, float) and not v.is_integer() and 1.0 <= v <= 3.0]
            if len(float_vals) >= 3:  # at least 3 float scores = it's the Internal row
                internal_score_row = idx
                break

    import sys
    print(f"====== FINAL internal_score_row: {internal_score_row} ======", file=sys.stderr, flush=True)
    print(f"====== tier_table_start: {tier_table_start} ======", file=sys.stderr, flush=True)
    # Save group weights to config_kv
    group_cfg = {'Materiality': 'materiality_weight',
                 'Criticality': 'criticality_weight',
                 'Complexity':  'complexity_weight'}
    for grp, w in group_weight_rows.items():
        db.execute('INSERT OR REPLACE INTO config_kv (key, value) VALUES (?,?)',
                   (group_cfg[grp], str(w)))
    db.commit()

    # Save param weights — multi-strategy to handle typos and name variations
    all_params_in_db = db.execute('SELECT id, sub_parameter FROM parameters').fetchall()
    
    for nk, w in param_weight_rows.items():
        updated = False
        
        # Strategy 1: exact param_map_norm lookup
        pid = param_map_norm.get(nk)
        if pid:
            db.execute('UPDATE parameters SET weight=? WHERE id=?', (w, pid))
            updated = True
        
        if not updated:
            # Strategy 2: direct SQL normalised match
            db.execute(
                'UPDATE parameters SET weight=? WHERE LOWER(TRIM(sub_parameter))=?',
                (w, nk)
            )
            updated = True

        if not updated:
            # Strategy 3: fuzzy — match first 6 chars (handles typos like Dependancy/Dependency)
            for p in all_params_in_db:
                db_nk = _norm(p['sub_parameter'])
                if len(nk) >= 6 and len(db_nk) >= 6 and nk[:6] == db_nk[:6]:
                    db.execute('UPDATE parameters SET weight=? WHERE id=?', (w, p['id']))
                    break

    db.commit()

    # Save individual scores per model per param
    def _get_param_id(sub_val):
        if sub_val in param_map:
            return param_map[sub_val]
        return param_map_norm.get(_norm(sub_val))

    for nk, row_idx in param_score_rows.items():
        row    = rows[row_idx]
        sub_raw = str(rows[row_idx][SUB_COL - 1] or '').strip()
        pid    = _get_param_id(sub_raw) or param_map_norm.get(nk)
        if not pid:
            continue
        for col_idx, model_id in model_cols:
            arr_idx = col_idx - 1
            val = row[arr_idx] if arr_idx < len(row) else None
            if val is None:
                continue
            try:
                level = max(1, min(3, int(val)))
            except (ValueError, TypeError):
                level = 1
            db.execute(
                'INSERT OR REPLACE INTO model_scores (model_id, parameter_id, level) VALUES (?,?,?)',
                (model_id, pid, level)
            )
            results['scores'] += 1
    db.commit()

    # Save computed scores from Internal row (row 22)
    if internal_score_row is not None:
        row = rows[internal_score_row]
        for col_idx, model_id in model_cols:
            arr_idx = col_idx - 1
            val = row[arr_idx] if arr_idx < len(row) else None
            if val is None:
                continue
            try:
                db.execute(
                    'UPDATE models SET computed_score=?, last_computed_at=datetime("now") WHERE id=?',
                    (round(float(val), 2), model_id)
                )
            except (ValueError, TypeError):
                pass
        db.commit()
        results['score_row_found'] = True

    # Load tier thresholds
    if tier_table_start is not None:
        db.execute('DELETE FROM tiers')
        sort_order = 0
        for ridx in range(tier_table_start + 1, min(tier_table_start + 8, len(rows))):
            row = rows[ridx]
            if not any(row):
                continue
            tier_name = str(row[0]).strip() if row[0] else None
            if not tier_name or tier_name.lower() in ('none', '', 'lower range', 'lower_range', 'tier'):
                continue
            try:
                lo = float(row[1]) if len(row) > 1 and row[1] is not None else 0.0
                hi = float(row[2]) if len(row) > 2 and row[2] is not None else 3.0
            except (ValueError, TypeError):
                continue
            db.execute(
                'INSERT INTO tiers (name, lower_bound, upper_bound, sort_order) VALUES (?,?,?,?)',
                (tier_name, lo, hi, sort_order)
            )
            results['tiers'] += 1
            sort_order += 1
        db.commit()


# ─── Format A: Standard ──────────────────────────────────────────────────────

def _load_standard_format(wb, results):
    db = get_db()

    ws = _sheet(wb, 'Models')
    if ws:
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            hdrs = _headers(rows[0])
            for row in rows[1:]:
                if not any(row):
                    continue
                d    = dict(zip(hdrs, row))
                name = d.get('name') or d.get('model_name')
                if not name:
                    continue
                if not db.execute('SELECT id FROM models WHERE name=?', (str(name),)).fetchone():
                    db.execute('INSERT INTO models (name, risk_type) VALUES (?,?)',
                               (str(name), str(d.get('risk_type') or '')))
                    results['models'] += 1
            db.commit()

    ws = _sheet(wb, 'Parameters')
    if ws:
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            hdrs = _headers(rows[0])
            for row in rows[1:]:
                if not any(row):
                    continue
                d   = dict(zip(hdrs, row))
                grp = d.get('group') or d.get('grp')
                if not grp:
                    continue
                db.execute(
                    'INSERT INTO parameters '
                    '(grp, sub_parameter, criteria, description, weight,'
                    ' level1_label, level2_label, level3_label) VALUES (?,?,?,?,?,?,?,?)',
                    (str(grp), str(d.get('sub_parameter') or ''),
                     str(d.get('criteria') or ''), str(d.get('description') or ''),
                     float(d.get('weight') or 1.0),
                     str(d.get('low_(1)') or 'Low'),
                     str(d.get('medium_(2)') or 'Medium'),
                     str(d.get('high_(3)') or 'High'))
                )
                results['parameters'] += 1
            db.commit()

    ws = _sheet(wb, 'Tiers')
    if ws:
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            hdrs = _headers(rows[0])
            db.execute('DELETE FROM tiers')
            for i, row in enumerate(rows[1:]):
                if not any(row):
                    continue
                d    = dict(zip(hdrs, row))
                name = d.get('name') or d.get('tier')
                if not name:
                    continue
                db.execute(
                    'INSERT INTO tiers (name, lower_bound, upper_bound, sort_order) VALUES (?,?,?,?)',
                    (str(name), float(d.get('lower_bound') or 0),
                     float(d.get('upper_bound') or 100), i)
                )
                results['tiers'] += 1
            db.commit()

    ws = _sheet(wb, 'Settings')
    if ws:
        for row in ws.iter_rows(values_only=True):
            if not row or not row[0]:
                continue
            key   = str(row[0]).strip()
            value = str(row[1]).strip() if len(row) > 1 and row[1] is not None else ''
            db.execute('INSERT OR REPLACE INTO config_kv (key, value) VALUES (?,?)', (key, value))
            results['settings'] += 1
        db.commit()

    ws = _sheet(wb, 'Scores')
    if ws:
        rows = list(ws.iter_rows(values_only=True))
        if rows:
            hdrs = _headers(rows[0])
            for row in rows[1:]:
                if not any(row):
                    continue
                d          = dict(zip(hdrs, row))
                model_name = d.get('model_name') or d.get('model')
                param_name = d.get('sub_parameter') or d.get('parameter')
                if not model_name or not param_name:
                    continue
                mr = db.execute('SELECT id FROM models WHERE name=?', (str(model_name),)).fetchone()
                pr = db.execute('SELECT id FROM parameters WHERE sub_parameter=?',
                                (str(param_name),)).fetchone()
                if not mr or not pr:
                    continue
                level = max(1, min(3, int(d.get('level') or 1)))
                db.execute(
                    'INSERT OR REPLACE INTO model_scores (model_id, parameter_id, level) VALUES (?,?,?)',
                    (mr['id'], pr['id'], level)
                )
                results['scores'] += 1
            db.commit()


def _debug_matrix(filepath):
    """Return a short diagnostic string about what the loader can see in the file."""
    wb = load_workbook(filepath, data_only=True)
    sheet_names = wb.sheetnames

    ws = None
    for sname in sheet_names:
        if sname.lower() in ('tiering_method', 'tiering method'):
            ws = wb[sname]
            break
    if ws is None:
        return f"No Tiering_Method sheet found. Sheets: {sheet_names}"

    rows = list(ws.iter_rows(values_only=True))
    col_a_vals = []
    internal_found = False
    tier_found = False
    for idx, row in enumerate(rows):
        if not any(row):
            continue
        v = str(row[0] or '').strip()
        if v:
            col_a_vals.append(f"row{idx+1}='{v}'")
        # Check any col in first 4 for 'internal'
        any_cols = [str(row[c] or '').strip().lower() for c in range(min(4, len(row)))]
        if any(c in ('internal', 'internal score') or c.startswith('internal') for c in any_cols if c):
            internal_found = True
        sub = str(row[1] or '').strip().lower() if len(row) > 1 else ''
        if v.lower() == 'tier' and sub in ('', 'lower range', 'lower_range'):
            tier_found = True

    # Count model cols in header
    header = rows[0] if rows else []
    non_empty_header = sum(1 for v in header if v)

    return (f"Sheets={sheet_names} | "
            f"Rows={len(rows)} | "
            f"Header non-empty cols={non_empty_header} | "
            f"Internal row found={internal_found} | "
            f"Tier table found={tier_found} | "
            f"Col-A values: {col_a_vals}")
